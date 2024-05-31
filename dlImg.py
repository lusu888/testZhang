from openpyxl import load_workbook
import requests
from tqdm import tqdm
import time

def dlImg(imgUrl,imageName):
    headers = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:60.0) Gecko/20100101 Firefox/60.0",
           "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
           "Accept-Language": "en-US,en;q=0.9"
           }
    img_data = requests.get(url=imgUrl, headers=headers).content
    with open( "img/" + imageName + '.png', 'wb') as handler:
        handler.write(img_data)




def main():
    print("Project Begin")
    excelFileName = "picpath.xlsx"

    
    workbook = load_workbook(filename=excelFileName)
    sheetObj = workbook.active

    i = 0

    productImgPath = sheetObj.cell(row = i+2, column = 2)
    sheetObj.cell(row=i+1, column = 3).value = "file Name" 

    maxRow = sheetObj.max_row - 1

    for i in tqdm(range(maxRow)):
        imageName = "Img_%d" % (i+2)
        dlImg(productImgPath.value,imageName)
        sheetObj.cell(row=i+2, column = 3).value = imageName +".png" 
        
        i += 1
        productImgPath = sheetObj.cell(row = i+2, column = 2)

    workbook.save(filename=excelFileName)



if __name__ == "__main__":
    main()

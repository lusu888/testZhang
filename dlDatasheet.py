from openpyxl import load_workbook
import requests
from tqdm import tqdm
import time
from selenium import webdriver
import os
import shutil
from selenium.webdriver.chrome.options import Options
import traceback

projectDir = "D:\\datasheetDl\\"
datasheetDlPath = projectDir + "tempfile\\" 
datasheetArchPath = projectDir + "datasheetAr\\"


def browserLogin(driver):
    """
    Input: webdriver.Chrome
    Output: webdriver.Chrome
    description: when the user has not login, or timeout, the browser will login.
    authour: Zhang Zhang 408131
    create date: 2024-03-28
    modify date: 2024-03-28
    """
    

    driver.switch_to.window(driver.window_handles[0])

    driver.get("https://it-b2b-plus.dahuasecurity.com/#/login")
    time.sleep(3)
    driver.maximize_window()
    time.sleep(2)

    userInput = driver.find_element("xpath","//input[@placeholder='User']")
    userInput.send_keys("408131")
    time.sleep(2)

    pwdInput = driver.find_element("xpath","//input[@placeholder='Password']")
    pwdInput.send_keys("Zz150317.2312")
    time.sleep(1)

    sbmBtn = driver.find_element("xpath","//button")
    sbmBtn.click()
    time.sleep(5)

    return driver






def dlDatasheetAction(driver, metelCode):

    #empty the dldir
    shutil.rmtree(datasheetDlPath)
    os.mkdir(datasheetDlPath)

    driver.switch_to.window(driver.window_handles[0])
    
    errInfo = ""
    driver.get("https://it-b2b-plus.dahuasecurity.com/#/Product/ProductList")
    time.sleep(3)


    driver.find_element("xpath","//Button[span[text()='Reset']]").click()
    time.sleep(3)


    pnInput = driver.find_element("xpath","//*[@id='app']/section/section/main/div[2]/section/main/div/div[1]/div[6]/div/div/div[2]/div/input")
    pnInput.send_keys(metelCode)
    time.sleep(2)
    driver.find_element("xpath","//ul/li/span[text()='"+metelCode+"']").click()
    time.sleep(3)
    driver.find_element("xpath","//Button[span[text()='Search']]").click()
    time.sleep(3)
 


    try:
        driver.find_element("xpath","//table/tbody/tr[1]/td[3]/div/span").click()
    except:

        driver.find_element("xpath","//*[@id='app']/section/section/main/div[2]/section/main/div/div[1]/div[7]/div/div/div[2]/div/input").click()
        time.sleep(2)

        driver.find_element("xpath","//ul/li/span[text()='All']").click()
        time.sleep(3)
        driver.find_element("xpath","//Button[span[text()='Search']]").click()
        time.sleep(3)
         

        try:
            driver.find_element("xpath","//table/tbody/tr[1]/td[3]/div/span").click()
        except:
            return False, "",driver,"no product"
    time.sleep(4)
    driver.switch_to.window(driver.window_handles[1])

    time.sleep(2)
    driver.find_element("xpath","//div[text()='Document']").click()
    
    time.sleep(2)
    datasheetCell = driver.find_element("xpath","//*[@id='pane-5']/div/div[1]/div/div[3]/table/tbody/tr[1]/td[2]/div")
    print(datasheetCell.text)


    time.sleep(2)
    datasheetCell.click()
    time.sleep(2)

    if int(datasheetCell.text) == 0:
        return False,"",driver, "no document"



    try:

        liElements = driver.find_elements("xpath", "//form/ul/li")

        time.sleep(2)
        isFindDatasheet = False

        for liElement in liElements:

            datasheetFileEle = liElement.find_element("xpath","span[2]")
            fileWebName = datasheetFileEle.text.lower()
            time.sleep(2)
            print(fileWebName)
            if "data" in fileWebName or "tech" in fileWebName:
                if fileWebName.endswith("zip") or fileWebName.endswith("indd"):
                    continue
                liElement.find_element("xpath","span[4]/button[1]/i").click()
                time.sleep(5)
                isFindDatasheet = True
                print("find datasheet")
                break
        if isFindDatasheet == False:
            for liElement in liElements:
                datasheetFileEle = liElement.find_element("xpath","span[2]")
                fileWebName = datasheetFileEle.text.lower()
                time.sleep(2)
                print(fileWebName)
                if fileWebName.endswith("pdf"):
                    liElement.find_element("xpath","span[4]/button[1]/i").click()
                    time.sleep(5)
                    isFindDatasheet = True
                    print("find datasheet")
                    break
                
        if isFindDatasheet == False:
            print("not sure to find datasheet")
            """

            datasheetFileEle = driver.find_element("xpath","//form/ul/li[1]/span[2]")
            time.sleep(3)
            driver.find_element("xpath","//form/ul/li[1]/span[4]/button[1]/i").click()
            time.sleep(5)
            """
            for liElement in liElements:
                datasheetFileEle = liElement.find_element("xpath","span[2]")
                fileWebName = datasheetFileEle.text.lower()
                time.sleep(2)
                print(fileWebName)
                liElement.find_element("xpath","span[4]/button[1]/i").click()
                time.sleep(5)


            

        """ 
        if int(datasheetCell.text) == 1:
            datasheetFileEle = driver.find_element("xpath","//form/ul/li[1]/span[2]")
            time.sleep(3)
            driver.find_element("xpath","//form/ul/li[1]/span[4]/button[1]/i").click()
            time.sleep(5)
        else:

            

            datasheetFileEle = driver.find_element("xpath","//form/ul/li[2]/span[2]")
            time.sleep(3)
            driver.find_element("xpath","//form/ul/li[2]/span[4]/button[1]/i").click()
            time.sleep(5)
        """
    except:
        return False,"",driver, "no document"
    fileNameDH = datasheetFileEle.text +"_"+metelCode + ".pdf"
    print(datasheetFileEle.text)
    
    
    driver.close()

    try:

        # dlFileName = os.listdir(datasheetDlPath)[0]
        for dlFileName in os.listdir(datasheetDlPath):
            print(dlFileName)
            if "pdf" in dlFileName:
                print("find pdf document")
                break
            else:
                os.remove(datasheetDlPath+dlFileName)
    except:
        return False, "",driver,"file error"
    if not "pdf" in dlFileName:
        print("not find pdf file")
        time.sleep(2)
        return False,"",driver, "need check document"
    shutil.move(datasheetDlPath+dlFileName,datasheetArchPath + fileNameDH)

    driver.switch_to.window(driver.window_handles[0])

    time.sleep(5)
    
    return True,fileNameDH, driver, errInfo
   
def dlFileFromDahuaSecu(driver, modelCode):
    errInfo = ""
    shutil.rmtree(datasheetDlPath)
    os.mkdir(datasheetDlPath)

    
    errInfo = ""
    driver.get("https://www.dahuasecurity.com/it/search/products?keyword=" + modelCode)
    time.sleep(3)


    productCount = int(driver.find_element("xpath","//span[@class='search-text'][text()='Prodotti']//following-sibling::span[@class='search-count']").text)
    time.sleep(3)

    if productCount == 0:
        print("no document")
        return False,"",driver, "no document 2"
    print("find document")


    dlElement = driver.find_element("xpath","//div[@class='search-result']/div[1]/div[2]/div[3]/div[2]/div[1]/a")
    time.sleep(2)
    fileWebName = dlElement.text.strip()
    fileNameDH = fileWebName +"_"+modelCode + ".pdf"

    print(fileWebName)
 
    dlElement.click()
    time.sleep(5)

    try:

        dlFileName = os.listdir(datasheetDlPath)[0]
        """
        for dlFileName in os.listdir(datasheetDlPath):
            print(dlFileName)
            if "pdf" in dlFileName:
                print("find pdf document")
                break
            else:
                os.remove(datasheetDlPath+dlFileName)
        """
    except:
        return False, "",driver,"file error"
    if not "pdf" in dlFileName:
        print("not find pdf file")
        time.sleep(2)
        return False,"",driver, "need check document"

    print("find pdf document")
    fileNameDH = fileNameDH.replace("/","_")
    shutil.move(datasheetDlPath+dlFileName,datasheetArchPath + fileNameDH)


    time.sleep(5)
 

    return True,fileNameDH, driver, errInfo



def main():
    print("Project Begin")

    if not os.path.exists(projectDir):
        os.makedirs(projectDir)
    if not os.path.exists(datasheetDlPath):
        os.makedirs(datasheetDlPath)
    if not os.path.exists(datasheetArchPath):
        os.makedirs(datasheetArchPath)

    options = Options()
    options.add_experimental_option('prefs', {
        "plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}], # Disable Chrome's PDF Viewer
        "download.default_directory": datasheetDlPath , 
        "download.extensions_to_open": "applications/pdf",
         "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True
        })    

    driver = webdriver.Chrome(options=options)

    


    
    driver = browserLogin(driver)



    excelFileName = "techsheet2.xlsx"

 
    workbook = load_workbook(filename=excelFileName)
    sheetObj = workbook.active

    i = 0
    startLine = 2
    workLine = 0

    maxRow = sheetObj.max_row - startLine + 1
    dhPartNr = ""

    
    for i in tqdm(range(maxRow)):
        workLine = startLine + i
        dhPartNr = sheetObj.cell(row=workLine, column = 2).value
        commentsStr = sheetObj.cell(row=workLine, column =6).value
        if dhPartNr == "#N/A":
            #or commentsStr != "input error" :
            continue
        print(dhPartNr)
        metelCode = sheetObj.cell(row=workLine, column =3).value
        modelCode = sheetObj.cell(row=workLine, column =4).value
        
        try: 
            dlOK, dhFileName, driver, errInfo = dlDatasheetAction(driver,metelCode )
            #dlOK, dhFileName, driver, errInfo = dlFileFromDahuaSecu(driver, modelCode)
            if dlOK:
                sheetObj.cell(row=workLine,column =5).value = dhFileName

                sheetObj.cell(row=workLine,column =6).value = ""
            else:
                sheetObj.cell(row=workLine,column =6).value = errInfo


        except Exception as error:
            print("login exception occurred: ", error)
            print(traceback.format_exc())
            
            try:

                driver = browserLogin(driver)
                dlOK, dhFileName, driver, errInfo = dlDatasheetAction(driver,metelCode )

                # dlOK, dhFileName, driver, errInfo = dlFileFromDahuaSecu(driver, modelCode)
                if dlOK:
                    sheetObj.cell(row=workLine,column =5).value = dhFileName

                    sheetObj.cell(row=workLine,column =6).value = ""
                else:
                    sheetObj.cell(row=workLine,column =6).value = errInfo
            except Exception as error:
                print("an exception occurred: ", error)
                print(traceback.format_exc())
     

                # break
            continue
        


    driver.quit

    workbook.save(filename=excelFileName)


if __name__ == "__main__":
    main()

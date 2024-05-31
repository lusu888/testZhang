from PIL import ImageFile, Image
Image.MAX_IMAGE_PIXELS = None
ImageFile.LOAD_TRUNCATED_IMAGES = True
from openpyxl import load_workbook
from tqdm import tqdm




def main():
    print("lower image begin")    
    excelFileName = "picpath.xlsx"

    workbook = load_workbook(filename=excelFileName)
    sheetObj = workbook.active

    maxRow = sheetObj.max_row - 1

    for i in tqdm(range(maxRow)):
        productImgName = sheetObj.cell(row = i+2, column = 3).value

        img = Image.open("img/" + productImgName)
        if img.mode != "RGB":
            img = img.convert("RGB")
        
        widthZ, heightZ = img.size
        widthZN, heightZN = widthZ, heightZ
        if widthZ > 1024:
            rateZ = widthZ/1024.0
            widthZN = 1024
            heightZN = int(heightZ/rateZ)
            """rateZ = widthZ/widthZN = heightZ/heightZN"""
            img.thumbnail((widthZN,heightZN))
        img.save("imglower/" + productImgName,"png", optimize=True)


if __name__ == "__main__":
    main()
